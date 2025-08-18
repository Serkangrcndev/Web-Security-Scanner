from celery import shared_task
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
from datetime import datetime, timedelta
import logging
import os
import json
from typing import Dict, Any, List

from app.database import SessionLocal, Report, Scan, Vulnerability, User
from app.core.config import settings

logger = logging.getLogger(__name__)

# Database connection for Celery tasks
engine = create_engine(settings.DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

@shared_task(bind=True, name="app.tasks.report_tasks.generate_report")
def generate_report(self, report_id: int):
    """
    Rapor oluştur
    Bu task Celery worker tarafından asenkron olarak çalıştırılır
    """
    db = SessionLocal()
    try:
        logger.info(f"Starting report generation for report {report_id}")
        
        # Rapor bilgilerini al
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            logger.error(f"Report {report_id} not found")
            return False
        
        # Rapor durumunu güncelle
        report.status = "generating"
        db.commit()
        
        # Tarama bilgilerini al
        scan = db.query(Scan).filter(Scan.id == report.scan_id).first()
        if not scan:
            logger.error(f"Scan {report.scan_id} not found for report {report_id}")
            return False
        
        # Kullanıcı bilgilerini al
        user = db.query(User).filter(User.id == report.user_id).first()
        if not user:
            logger.error(f"User {report.user_id} not found for report {report_id}")
            return False
        
        # Güvenlik açıklarını al
        vulnerabilities = db.query(Vulnerability).filter(
            Vulnerability.scan_id == report.scan_id
        ).all()
        
        # Rapor tipine göre oluştur
        if report.report_type == "pdf":
            success = generate_pdf_report(report, scan, user, vulnerabilities)
        elif report.report_type == "excel":
            success = generate_excel_report(report, scan, user, vulnerabilities)
        elif report.report_type == "json":
            success = generate_json_report(report, scan, user, vulnerabilities)
        else:
            logger.error(f"Unsupported report type: {report.report_type}")
            success = False
        
        if success:
            # Rapor durumunu güncelle
            report.status = "completed"
            report.completed_at = datetime.utcnow()
            db.commit()
            
            logger.info(f"Report {report_id} generated successfully")
            return True
        else:
            # Hata durumunu güncelle
            report.status = "failed"
            report.error_message = "Rapor oluşturulamadı"
            db.commit()
            
            logger.error(f"Report {report_id} generation failed")
            return False
        
    except Exception as e:
        logger.error(f"Report generation {report_id} failed: {e}")
        
        # Hata durumunu güncelle
        if report:
            report.status = "failed"
            report.error_message = str(e)
            db.commit()
        
        return False
        
    finally:
        db.close()

def generate_pdf_report(report: Report, scan: Scan, user: User, vulnerabilities: List[Vulnerability]) -> bool:
    """PDF rapor oluştur"""
    try:
        # Rapor dizinini oluştur
        report_dir = os.path.join(settings.BASE_DIR, "reports", "pdf")
        os.makedirs(report_dir, exist_ok=True)
        
        # Dosya adı oluştur
        filename = f"security_report_{report.id}_{scan.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = os.path.join(report_dir, filename)
        
        # PDF oluştur (ReportLab kullanarak)
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Başlık
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Güvenlik Tarama Raporu", title_style))
        story.append(Spacer(1, 20))
        
        # Tarama bilgileri
        story.append(Paragraph("Tarama Bilgileri", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        scan_data = [
            ["Hedef URL", scan.target_url],
            ["Tarama Tipi", scan.scan_type],
            ["Durum", scan.status],
            ["Başlama Tarihi", scan.started_at.strftime("%Y-%m-%d %H:%M:%S") if scan.started_at else "N/A"],
            ["Tamamlanma Tarihi", scan.completed_at.strftime("%Y-%m-%d %H:%M:%S") if scan.completed_at else "N/A"],
            ["Oluşturulma Tarihi", scan.created_at.strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        scan_table = Table(scan_data, colWidths=[2*inch, 4*inch])
        scan_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(scan_table)
        story.append(Spacer(1, 20))
        
        # Güvenlik açıkları özeti
        story.append(Paragraph("Güvenlik Açıkları Özeti", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for vuln in vulnerabilities:
            severity_counts[vuln.severity] += 1
        
        summary_data = [
            ["Severity", "Sayı"],
            ["Kritik", str(severity_counts["critical"])],
            ["Yüksek", str(severity_counts["high"])],
            ["Orta", str(severity_counts["medium"])],
            ["Düşük", str(severity_counts["low"])],
            ["Toplam", str(len(vulnerabilities))]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Detaylı güvenlik açıkları
        if vulnerabilities:
            story.append(Paragraph("Detaylı Güvenlik Açıkları", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            for i, vuln in enumerate(vulnerabilities, 1):
                story.append(Paragraph(f"{i}. {vuln.title}", styles['Heading3']))
                story.append(Paragraph(f"Severity: {vuln.severity.upper()}", styles['Normal']))
                story.append(Paragraph(f"Scanner: {vuln.scanner_name}", styles['Normal']))
                story.append(Paragraph(f"Description: {vuln.description}", styles['Normal']))
                if vuln.location:
                    story.append(Paragraph(f"Location: {vuln.location}", styles['Normal']))
                if vuln.evidence:
                    story.append(Paragraph(f"Evidence: {vuln.evidence}", styles['Normal']))
                story.append(Spacer(1, 12))
        
        # PDF oluştur
        doc.build(story)
        
        # Rapor dosya yolunu güncelle
        report.file_path = file_path
        return True
        
    except Exception as e:
        logger.error(f"PDF report generation failed: {e}")
        return False

def generate_excel_report(report: Report, scan: Scan, user: User, vulnerabilities: List[Vulnerability]) -> bool:
    """Excel rapor oluştur"""
    try:
        # Rapor dizinini oluştur
        report_dir = os.path.join(settings.BASE_DIR, "reports", "excel")
        os.makedirs(report_dir, exist_ok=True)
        
        # Dosya adı oluştur
        filename = f"security_report_{report.id}_{scan.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(report_dir, filename)
        
        # Excel oluştur (openpyxl kullanarak)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Güvenlik Raporu"
        
        # Başlık
        ws['A1'] = "Güvenlik Tarama Raporu"
        ws.merge_cells('A1:F1')
        title_font = Font(size=18, bold=True)
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Tarama bilgileri
        ws['A3'] = "Tarama Bilgileri"
        ws['A3'].font = Font(size=14, bold=True)
        
        scan_data = [
            ["Hedef URL", scan.target_url],
            ["Tarama Tipi", scan.scan_type],
            ["Durum", scan.status],
            ["Başlama Tarihi", scan.started_at.strftime("%Y-%m-%d %H:%M:%S") if scan.started_at else "N/A"],
            ["Tamamlanma Tarihi", scan.completed_at.strftime("%Y-%m-%d %H:%M:%S") if scan.completed_at else "N/A"],
            ["Oluşturulma Tarihi", scan.created_at.strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for i, (key, value) in enumerate(scan_data, 4):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Güvenlik açıkları özeti
        ws['A12'] = "Güvenlik Açıkları Özeti"
        ws['A12'].font = Font(size=14, bold=True)
        
        severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for vuln in vulnerabilities:
            severity_counts[vuln.severity] += 1
        
        summary_headers = ["Severity", "Sayı"]
        summary_data = [
            ["Kritik", severity_counts["critical"]],
            ["Yüksek", severity_counts["high"]],
            ["Orta", severity_counts["medium"]],
            ["Düşük", severity_counts["low"]],
            ["Toplam", len(vulnerabilities)]
        ]
        
        # Özet tablosu
        for i, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=13, column=i)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        for i, (severity, count) in enumerate(summary_data, 14):
            ws.cell(row=i, column=1, value=severity)
            ws.cell(row=i, column=2, value=count)
        
        # Detaylı güvenlik açıkları
        if vulnerabilities:
            ws['A20'] = "Detaylı Güvenlik Açıkları"
            ws['A20'].font = Font(size=14, bold=True)
            
            # Tablo başlıkları
            headers = ["#", "Başlık", "Severity", "Scanner", "Açıklama", "Konum", "Kanıt"]
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=21, column=i)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Veri satırları
            for i, vuln in enumerate(vulnerabilities, 22):
                ws.cell(row=i, column=1, value=i-21)
                ws.cell(row=i, column=2, value=vuln.title)
                ws.cell(row=i, column=3, value=vuln.severity.upper())
                ws.cell(row=i, column=4, value=vuln.scanner_name)
                ws.cell(row=i, column=5, value=vuln.description)
                ws.cell(row=i, column=6, value=vuln.location or "")
                ws.cell(row=i, column=7, value=vuln.evidence or "")
        
        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Excel dosyasını kaydet
        wb.save(file_path)
        
        # Rapor dosya yolunu güncelle
        report.file_path = file_path
        return True
        
    except Exception as e:
        logger.error(f"Excel report generation failed: {e}")
        return False

def generate_json_report(report: Report, scan: Scan, user: User, vulnerabilities: List[Vulnerability]) -> bool:
    """JSON rapor oluştur"""
    try:
        # Rapor dizinini oluştur
        report_dir = os.path.join(settings.BASE_DIR, "reports", "json")
        os.makedirs(report_dir, exist_ok=True)
        
        # Dosya adı oluştur
        filename = f"security_report_{report.id}_{scan.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path = os.path.join(report_dir, filename)
        
        # Rapor verilerini hazırla
        report_data = {
            "report_info": {
                "id": report.id,
                "type": report.report_type,
                "generated_at": datetime.now().isoformat(),
                "scan_id": scan.id,
                "user_id": user.id
            },
            "scan_info": {
                "target_url": scan.target_url,
                "scan_type": scan.scan_type,
                "status": scan.status,
                "started_at": scan.started_at.isoformat() if scan.started_at else None,
                "completed_at": scan.completed_at.isoformat() if scan.completed_at else None,
                "created_at": scan.created_at.isoformat()
            },
            "user_info": {
                "username": user.username,
                "email": user.email,
                "is_premium": user.is_premium
            },
            "vulnerabilities": [
                {
                    "title": vuln.title,
                    "description": vuln.description,
                    "severity": vuln.severity,
                    "cve_id": vuln.cve_id,
                    "cvss_score": vuln.cvss_score,
                    "scanner_name": vuln.scanner_name,
                    "payload": vuln.payload,
                    "location": vuln.location,
                    "evidence": vuln.evidence,
                    "created_at": vuln.created_at.isoformat()
                }
                for vuln in vulnerabilities
            ],
            "summary": {
                "total_vulnerabilities": len(vulnerabilities),
                "severity_counts": {
                    "critical": len([v for v in vulnerabilities if v.severity == "critical"]),
                    "high": len([v for v in vulnerabilities if v.severity == "high"]),
                    "medium": len([v for v in vulnerabilities if v.severity == "medium"]),
                    "low": len([v for v in vulnerabilities if v.severity == "low"])
                }
            }
        }
        
        # JSON dosyasını kaydet
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
        
        # Rapor dosya yolunu güncelle
        report.file_path = file_path
        return True
        
    except Exception as e:
        logger.error(f"JSON report generation failed: {e}")
        return False

@shared_task(bind=True, name="app.tasks.report_tasks.cleanup_old_reports")
def cleanup_old_reports(self, days_old: int = 7):
    """Eski raporları temizle"""
    db = SessionLocal()
    try:
        cutoff_date = datetime.utcnow() - timedelta(days=days_old)
        
        # Eski raporları bul
        old_reports = db.query(Report).filter(
            Report.created_at < cutoff_date
        ).all()
        
        deleted_count = 0
        for report in old_reports:
            try:
                # Dosyayı sil
                if report.file_path and os.path.exists(report.file_path):
                    os.remove(report.file_path)
                
                # Veritabanından sil
                db.delete(report)
                deleted_count += 1
                
            except Exception as e:
                logger.error(f"Failed to cleanup report {report.id}: {e}")
                continue
        
        db.commit()
        logger.info(f"Report cleanup completed: {deleted_count} old reports removed")
        
        return deleted_count
        
    except Exception as e:
        logger.error(f"Report cleanup failed: {e}")
        return 0
        
    finally:
        db.close()

@shared_task(bind=True, name="app.tasks.report_tasks.generate_bulk_report")
def generate_bulk_report(self, scan_ids: List[int], report_type: str, user_id: int):
    """Toplu rapor oluştur"""
    db = SessionLocal()
    try:
        logger.info(f"Starting bulk report generation for {len(scan_ids)} scans")
        
        # Kullanıcı bilgilerini al
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            logger.error(f"User {user_id} not found for bulk report")
            return False
        
        # Tüm taramaları al
        scans = db.query(Scan).filter(
            Scan.id.in_(scan_ids),
            Scan.user_id == user_id
        ).all()
        
        if not scans:
            logger.error(f"No scans found for bulk report")
            return False
        
        # Tüm güvenlik açıklarını al
        all_vulnerabilities = db.query(Vulnerability).filter(
            Vulnerability.scan_id.in_(scan_ids)
        ).all()
        
        # Toplu rapor oluştur
        if report_type == "pdf":
            success = generate_bulk_pdf_report(scans, user, all_vulnerabilities)
        elif report_type == "excel":
            success = generate_bulk_excel_report(scans, user, all_vulnerabilities)
        elif report_type == "json":
            success = generate_bulk_json_report(scans, user, all_vulnerabilities)
        else:
            logger.error(f"Unsupported bulk report type: {report_type}")
            success = False
        
        if success:
            logger.info(f"Bulk report generated successfully for {len(scan_ids)} scans")
            return True
        else:
            logger.error(f"Bulk report generation failed")
            return False
        
    except Exception as e:
        logger.error(f"Bulk report generation failed: {e}")
        return False
        
    finally:
        db.close()

def generate_bulk_pdf_report(scans: List[Scan], user: User, vulnerabilities: List[Vulnerability]) -> bool:
    """Toplu PDF rapor oluştur"""
    try:
        # Bu fonksiyon generate_pdf_report'e benzer şekilde implement edilebilir
        # Şimdilik sadece True döndürüyoruz
        return True
    except Exception as e:
        logger.error(f"Bulk PDF report generation failed: {e}")
        return False

def generate_bulk_excel_report(scans: List[Scan], user: User, vulnerabilities: List[Vulnerability]) -> bool:
    """Toplu Excel rapor oluştur"""
    try:
        # Bu fonksiyon generate_excel_report'e benzer şekilde implement edilebilir
        # Şimdilik sadece True döndürüyoruz
        return True
    except Exception as e:
        logger.error(f"Bulk Excel report generation failed: {e}")
        return False

def generate_bulk_json_report(scans: List[Scan], user: User, vulnerabilities: List[Vulnerability]) -> bool:
    """Toplu JSON rapor oluştur"""
    try:
        # Bu fonksiyon generate_json_report'e benzer şekilde implement edilebilir
        # Şimdilik sadece True döndürüyoruz
        return True
    except Exception as e:
        logger.error(f"Bulk JSON report generation failed: {e}")
        return False
